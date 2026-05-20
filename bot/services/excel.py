import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS: list[tuple[str, str]] = [
    ("id",          "ID"),
    ("created_at",  "Дата регистрации"),
    ("fullname",    "ФИО"),
    ("birthdate",   "Дата рождения"),
    ("city",        "Город"),
    ("phone",       "Телефон"),
    ("email",       "Email"),
    ("telegram",    "Telegram"),
    ("vk",          "VK"),
    ("status",      "Статус"),
    ("education",   "Образование"),
    ("university",  "Университет"),
    ("faculty",     "Факультет"),
    ("course",      "Курс"),
    ("grad_year",   "Год выпуска"),
    ("track",       "Трек"),
]

_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_ROW_FONT    = Font(name="Arial", size=10)
_ALT_FILL    = PatternFill("solid", start_color="D6E4F0")

_COL_WIDTHS: dict[str, int] = {
    "id": 6, "created_at": 18, "fullname": 28, "birthdate": 14,
    "city": 16, "phone": 16, "email": 26, "telegram": 18,
    "vk": 18, "status": 14, "education": 16, "university": 30,
    "faculty": 26, "course": 8, "grad_year": 12, "track": 18,
}


def build_short_xlsx_bytes(rows: list[dict]) -> bytes:
    """3-column format for scheduled exports: Фамилия | Имя Отчество | Email."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    headers = ["Фамилия", "Имя Отчество", "Email"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        fullname = (row.get("fullname") or "").strip()
        parts = fullname.split(" ", 1)
        last_name = parts[0]
        first_middle = parts[1] if len(parts) > 1 else ""
        email = row.get("email") or ""

        for col_idx, val in enumerate([last_name, first_middle, email], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _ROW_FONT
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_xlsx_bytes(rows: list[dict]) -> bytes:
    keys   = [c[0] for c in COLUMNS]
    labels = [c[1] for c in COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    for row_idx, row in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(keys, start=1):
            val = row.get(key)
            if key == "created_at" and val:
                val = val[:19].replace("T", " ")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _ROW_FONT
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    for col_idx, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(key, 16)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
