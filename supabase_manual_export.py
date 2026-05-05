#!/usr/bin/env python3
"""
Выгрузка participants из Supabase → Excel-файл
Запуск: python export_participants.py
"""

import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

COLUMNS = [
    ("id",           "ID"),
    ("created_at",   "Дата регистрации"),
    ("fullname",     "ФИО"),
    ("birthdate",    "Дата рождения"),
    ("city",         "Город"),
    ("phone",        "Телефон"),
    ("email",        "Email"),
    ("telegram",     "Telegram"),
    ("vk",           "VK"),
    ("status",       "Статус"),
    ("education",    "Образование"),
    ("university",   "Университет"),
    ("faculty",      "Факультет"),
    ("course",       "Курс"),
    ("grad_year",    "Год выпуска"),
    ("track",        "Трек"),
]

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ROW_FONT    = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")


def fetch_participants() -> list[dict]:
    client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return client.table("participants").select("*").order("id").execute().data


def build_xlsx(rows: list[dict], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    keys   = [c[0] for c in COLUMNS]
    labels = [c[1] for c in COLUMNS]

    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(keys, start=1):
            val = row.get(key)
            if key == "created_at" and val:
                val = val[:19].replace("T", " ")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = ROW_FONT
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    col_widths = {
        "id": 6, "created_at": 18, "fullname": 28, "birthdate": 14,
        "city": 16, "phone": 16, "email": 26, "telegram": 18,
        "vk": 18, "status": 14, "education": 16, "university": 30,
        "faculty": 26, "course": 8, "grad_year": 12, "track": 18,
    }
    for col_idx, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 16)

    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    print("⏳ Получаю данные из Supabase...")
    rows = fetch_participants()
    print(f"   Найдено записей: {len(rows)}")

    filename = f"participants_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    print(f"⏳ Сохраняю {filename}...")
    build_xlsx(rows, filename)
    print(f"✅ Готово → {filename}")